"""
superadmin_monitoring.py — Monitoring natijalarini yuklab olish (superadmin uchun)

Oqim:
  /monitoring yoki tugma → Maktab tanlash → Guruh tanlash → Natijalar + Excel yuklash
"""
from __future__ import annotations

import io
from datetime import datetime

from aiogram import F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from school_bot.database.models import Group, School

router = Router(name="superadmin_monitoring")


class MonitoringStates(StatesGroup):
    choose_school = State()
    choose_group  = State()


# ── Helpers ─────────────────────────────────────────────────────────────────

async def _list_schools_with_results(session: AsyncSession) -> list[dict]:
    """Monitoring natijasi bor maktablarni qaytaradi."""
    try:
        rows = await session.execute(text("""
            SELECT DISTINCT
                s.id           AS school_id,
                s.name         AS school_name,
                COUNT(tr.id)   AS result_count
            FROM bot_schools s
            JOIN monitoring_testresult tr
                ON tr.school_id::text = s.alochi_school_id
            GROUP BY s.id, s.name
            ORDER BY s.name
        """))
        return [{"id": r.school_id, "name": r.school_name, "count": r.result_count} for r in rows]
    except Exception:
        # Fallback: just list all schools
        rows = await session.execute(select(School).order_by(School.name))
        return [{"id": s.id, "name": s.name, "count": 0} for s in rows.scalars()]


async def _list_groups_for_school(session: AsyncSession, school_id: int) -> list[dict]:
    """Maktab uchun guruhlarni (monitoring natijasi bor) qaytaradi."""
    try:
        rows = await session.execute(text("""
            SELECT DISTINCT
                g.id           AS group_id,
                g.name         AS group_name,
                g.alochi_group_id,
                COUNT(tr.id)   AS result_count
            FROM bot_groups g
            LEFT JOIN monitoring_testresult tr
                ON tr.group_id::text = g.alochi_group_id::text
            WHERE g.school_id = :sid
            GROUP BY g.id, g.name, g.alochi_group_id
            ORDER BY g.name
        """), {"sid": school_id})
        return [{"id": r.group_id, "name": r.group_name,
                 "alochi_id": str(r.alochi_group_id) if r.alochi_group_id else None,
                 "count": r.result_count} for r in rows]
    except Exception:
        rows = await session.execute(
            select(Group).where(Group.school_id == school_id).order_by(Group.name)
        )
        return [{"id": g.id, "name": g.name,
                 "alochi_id": str(g.alochi_group_id) if g.alochi_group_id else None,
                 "count": 0} for g in rows.scalars()]


async def _get_results(session: AsyncSession,
                       school_id: int | None = None,
                       group_alochi_id: str | None = None) -> list[dict]:
    """Monitoring natijalarini to'g'ridan DB dan oladi."""
    filters = []
    params: dict = {}

    if school_id is not None:
        filters.append("""
            tr.school_id IN (
                SELECT alochi_school_id::uuid FROM bot_schools WHERE id = :school_id
            )
        """)
        params["school_id"] = school_id

    if group_alochi_id:
        filters.append("tr.group_id = :group_id")
        params["group_id"] = group_alochi_id

    where = ("WHERE " + " AND ".join(filters)) if filters else ""

    try:
        rows = await session.execute(text(f"""
            SELECT
                tr.id,
                tr.created_at,
                tr.variant,
                tr.math_score,
                tr.eng_score,
                tr.total_pct,
                tr.passed,
                p.title        AS package_title,
                u.first_name   AS first_name,
                u.last_name    AS last_name,
                g.grade        AS grade
            FROM monitoring_testresult tr
            LEFT JOIN monitoring_monitoringpackage p ON p.id = tr.package_id
            LEFT JOIN monitoring_monitoringcred   c ON c.id = tr.cred_id
            LEFT JOIN monitoring_student          u ON u.id = c.student_id
            LEFT JOIN monitoring_studentgroup     g ON g.id = tr.group_id
            {where}
            ORDER BY tr.created_at DESC
            LIMIT 500
        """), params)
        return [dict(r._mapping) for r in rows]
    except Exception as exc:
        return []


def _build_excel(results: list[dict], title: str) -> bytes:
    """openpyxl bilan Excel fayl yaratadi."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # openpyxl yo'q bo'lsa CSV qaytaradi
        return _build_csv(results).encode("utf-8-sig")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    # Header
    headers = ["#", "Ism Familiya", "Sinf", "Variant",
               "Matematika", "Ingliz tili", "Jami %", "O'tdi/O'tmadi",
               "Paket", "Sana"]
    header_fill = PatternFill("solid", fgColor="F97316")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D4D4D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    # Data
    for i, r in enumerate(results, 1):
        name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip() or "—"
        grade = str(r.get("grade", "—"))
        passed_txt = "✅ O'tdi" if r.get("passed") else "❌ O'tmadi"
        sana = r.get("created_at")
        sana_str = sana.strftime("%d.%m.%Y %H:%M") if isinstance(sana, datetime) else str(sana or "—")

        row_data = [
            i, name, grade, r.get("variant", "—"),
            r.get("math_score", 0), r.get("eng_score", 0),
            f"{r.get('total_pct', 0)}%", passed_txt,
            r.get("package_title", "—"), sana_str,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                       vertical="center")
            cell.border = border
            if r.get("passed"):
                cell.fill = PatternFill("solid", fgColor="F0FDF4")
        ws.row_dimensions[i + 1].height = 18

    # Column widths
    widths = [5, 22, 8, 10, 14, 14, 10, 16, 26, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv(results: list[dict]) -> str:
    lines = ["#,Ism Familiya,Sinf,Variant,Matematika,Ingliz tili,Jami %,O'tdi/O'tmadi,Paket,Sana"]
    for i, r in enumerate(results, 1):
        name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip() or "—"
        passed = "O'tdi" if r.get("passed") else "O'tmadi"
        sana = r.get("created_at")
        sana_str = sana.strftime("%d.%m.%Y %H:%M") if isinstance(sana, datetime) else str(sana or "—")
        lines.append(
            f"{i},{name},{r.get('grade','—')},{r.get('variant','—')},"
            f"{r.get('math_score',0)},{r.get('eng_score',0)},"
            f"{r.get('total_pct',0)}%,{passed},"
            f"{r.get('package_title','—')},{sana_str}"
        )
    return "\n".join(lines)


def _schools_keyboard(schools: list[dict]) -> InlineKeyboardMarkup:
    rows = []
    for s in schools[:24]:
        cnt = f" ({s['count']})" if s.get("count") else ""
        rows.append([InlineKeyboardButton(
            text=f"🏫 {s['name']}{cnt}",
            callback_data=f"mon_sch:{s['id']}",
        )])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="mon_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _groups_keyboard(groups: list[dict], school_id: int) -> InlineKeyboardMarkup:
    rows = []
    # "Hammasi" button
    rows.append([InlineKeyboardButton(
        text="📥 Barcha guruhlar (hammasi)",
        callback_data=f"mon_grp:{school_id}:ALL",
    )])
    for g in groups[:20]:
        cnt = f" ({g['count']})" if g.get("count") else ""
        rows.append([InlineKeyboardButton(
            text=f"👥 {g['name']}{cnt}",
            callback_data=f"mon_grp:{school_id}:{g['alochi_id'] or g['id']}",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="mon_back")])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="mon_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ── Entry points ─────────────────────────────────────────────────────────────

@router.message(Command("monitoring"))
async def monitoring_cmd(message: Message, session: AsyncSession,
                          state: FSMContext, is_superadmin: bool = False) -> None:
    if not is_superadmin:
        await message.answer("⛔ Ruxsat yo'q.")
        return
    await _show_schools(message, session, state)


async def _show_schools(target: Message | CallbackQuery,
                        session: AsyncSession, state: FSMContext) -> None:
    schools = await _list_schools_with_results(session)
    if not schools:
        text_msg = "📭 Hozircha monitoring natijalari yo'q."
        if isinstance(target, CallbackQuery):
            await target.message.edit_text(text_msg)
        else:
            await target.answer(text_msg)
        return

    await state.set_state(MonitoringStates.choose_school)
    await state.update_data(schools={str(s["id"]): s for s in schools})

    kb = _schools_keyboard(schools)
    text_msg = "🏫 <b>Maktabni tanlang:</b>\n<i>Monitoring natijalarini yuklab olish uchun</i>"
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text_msg, reply_markup=kb, parse_mode="HTML")
        await target.answer()
    else:
        await target.answer(text_msg, reply_markup=kb, parse_mode="HTML")


@router.callback_query(MonitoringStates.choose_school, F.data.startswith("mon_sch:"))
async def school_selected(callback: CallbackQuery, session: AsyncSession,
                           state: FSMContext, is_superadmin: bool = False) -> None:
    if not is_superadmin:
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return

    school_id = int(callback.data.split(":")[1])
    data = await state.get_data()
    schools = data.get("schools", {})
    school_name = schools.get(str(school_id), {}).get("name", f"#{school_id}")

    groups = await _list_groups_for_school(session, school_id)
    if not groups:
        await callback.message.edit_text(
            f"📭 <b>{school_name}</b> maktabida guruh topilmadi.",
            parse_mode="HTML",
        )
        await callback.answer()
        return

    await state.set_state(MonitoringStates.choose_group)
    await state.update_data(school_id=school_id, school_name=school_name)

    kb = _groups_keyboard(groups, school_id)
    await callback.message.edit_text(
        f"🏫 <b>{school_name}</b>\n👥 <b>Guruhni tanlang:</b>",
        reply_markup=kb, parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(MonitoringStates.choose_group, F.data.startswith("mon_grp:"))
async def group_selected(callback: CallbackQuery, session: AsyncSession,
                          state: FSMContext, is_superadmin: bool = False) -> None:
    if not is_superadmin:
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return

    parts = callback.data.split(":")
    school_id = int(parts[1])
    group_key = parts[2]  # UUID or "ALL"

    data = await state.get_data()
    school_name = data.get("school_name", "—")

    await callback.message.edit_text("⏳ Natijalar yuklanmoqda...")
    await callback.answer()

    group_id = None if group_key == "ALL" else group_key
    group_label = "Barcha guruhlar" if group_key == "ALL" else group_key

    results = await _get_results(
        session,
        school_id=school_id,
        group_alochi_id=group_id,
    )

    if not results:
        await callback.message.edit_text(
            f"📭 <b>{school_name}</b> — {group_label}\nNatijalar topilmadi.",
            parse_mode="HTML",
        )
        await state.clear()
        return

    # Summary message
    passed = sum(1 for r in results if r.get("passed"))
    avg_pct = int(sum(r.get("total_pct", 0) for r in results) / len(results))
    summary = (
        f"📊 <b>{school_name}</b>\n"
        f"👥 {group_label}\n"
        f"━━━━━━━━━━━\n"
        f"Jami: <b>{len(results)}</b> ta natija\n"
        f"O'tdi: <b>{passed}</b> / {len(results)}\n"
        f"O'rtacha: <b>{avg_pct}%</b>\n"
        f"━━━━━━━━━━━\n"
        f"📥 Excel fayl tayyorlanmoqda..."
    )
    await callback.message.edit_text(summary, parse_mode="HTML")

    # Excel file
    filename = f"natijalar_{school_name[:15]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    excel_bytes = _build_excel(results, school_name)

    await callback.message.answer_document(
        document=BufferedInputFile(excel_bytes, filename=filename),
        caption=(
            f"📊 <b>{school_name}</b> — {group_label}\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
            f"📝 {len(results)} ta natija"
        ),
        parse_mode="HTML",
    )
    await state.clear()


@router.callback_query(F.data == "mon_back")
async def monitoring_back(callback: CallbackQuery, session: AsyncSession,
                           state: FSMContext, is_superadmin: bool = False) -> None:
    if not is_superadmin:
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return
    await _show_schools(callback, session, state)


@router.callback_query(F.data == "mon_cancel")
async def monitoring_cancel(callback: CallbackQuery, state: FSMContext,
                             is_superadmin: bool = False) -> None:
    await state.clear()
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.answer()


# ── Keyboard button handler ───────────────────────────────────────────────────

@router.message(F.text == "📋 Monitoring Natijalar")
async def monitoring_btn_handler(message: Message, session: AsyncSession,
                                  state: FSMContext, is_superadmin: bool = False) -> None:
    if not is_superadmin:
        return
    await _show_schools(message, session, state)
